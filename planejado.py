import os  # Importa o módulo os para interagir com o sistema de arquivos
import time  # Importa o módulo time para manipular o tempo, como pausas no código
from datetime import datetime  # Importa o módulo datetime para manipulação de data e hora
import pandas as pd  # Importa a biblioteca pandas para manipulação de dados em tabelas (DataFrames)
from selenium import webdriver  # Importa a biblioteca webdriver do Selenium para controle do navegador
from selenium.webdriver.edge.service import Service as EdgeService  # Importa a classe Service para controlar o serviço do navegador Edge
from selenium.webdriver.edge.options import Options  # Importa a classe Options para configurar opções do navegador Edge
from selenium.webdriver.common.by import By  # Importa a classe By para selecionar elementos no DOM via diferentes métodos
from openpyxl import load_workbook  # Importa a função load_workbook do openpyxl para carregar arquivos Excel
from openpyxl.styles import Font, PatternFill, Border, Side  # Importa classes para aplicar estilos no Excel
import config  # Importa o arquivo de configurações para acessar credenciais e caminhos do sistema

# Função para configurar o driver do navegador
def configurar_driver():
    edge_options = Options()  # Cria uma instância de opções para o navegador Edge
    edge_service = EdgeService(executable_path=config.EDGE_DRIVER_PATH)  # Configura o caminho do driver do Edge a partir do arquivo de configurações
    driver = webdriver.Edge(service=edge_service, options=edge_options)  # Inicializa o driver do navegador Edge com as opções e serviço configurados
    driver.maximize_window()  # Maximiza a janela do navegador
    return driver  # Retorna o driver configurado

# Função para realizar o login no sistema
def realizar_login(driver, username, password):
    driver.get('https://suporte.santamarcelinacultura.org.br/planejamento/front/central.php')  # Acessa a URL do sistema de login
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/form/div/div[1]/div[2]/input').send_keys(username)  # Preenche o campo de usuário
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/form/div/div[1]/div[3]/input').send_keys(password)  # Preenche o campo de senha
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/form/div/div[1]/div[5]/button').click()  # Clica no botão de login
    time.sleep(3)  # Aguarda 3 segundos para o login ser processado

# Função para navegar até o painel e realizar ações
def navegar_painel(driver, username, password):
    driver.find_element(By.XPATH, '/html/body/div[2]/header/div/div[2]/ul/li[1]/a/span').click()  # Clica em ativos
    time.sleep(2)  # Aguarda 2 segundos
    driver.find_element(By.XPATH, '/html/body/div[2]/header/div/div[2]/ul/li[1]/div/div/div[2]/a[5]').click()  # clica em painel
    time.sleep(2)  # Aguarda 2 segundos
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/main/table/tbody/tr[2]/td/div[1]/a').click()  # Clica em acessar painel
    time.sleep(3)  # Aguarda 3 segundos
    driver.find_element(By.XPATH, '/html/body/div/div/div[2]/form/div[1]/input').send_keys(username)  # Preenche o campo de usuário
    driver.find_element(By.XPATH, '/html/body/div/div/div[2]/form/div[2]/input').send_keys(password)  # Preenche o campo de senha
    driver.find_element(By.XPATH, '/html/body/div/div/div[2]/form/button').click()  # Clica para autenticar novamente
    time.sleep(3)  # Aguarda 3 segundos
    driver.find_element(By.XPATH, '/html/body/nav/button[4]').click()  # Clica em um botão para avançar no processo

# Função para adicionar bordas em todas as células de uma planilha
def adicionar_bordas(sheet):
    thin_border = Border(  # Cria uma borda fina para as células
        left=Side(border_style="thin", color="000000"),  # Define a borda esquerda
        right=Side(border_style="thin", color="000000"),  # Define a borda direita
        top=Side(border_style="thin", color="000000"),  # Define a borda superior
        bottom=Side(border_style="thin", color="000000")  # Define a borda inferior
    )
    
    for row in sheet.iter_rows():  # Itera sobre todas as linhas da planilha
        for cell in row:  # Itera sobre todas as células da linha
            cell.border = thin_border  # Aplica a borda fina a cada célula

# Função para renomear, formatar e adicionar a coluna 'AREA2' ao arquivo
def renomear_formatar_arquivo():
    download_dir = os.path.join(os.path.expanduser('~'), 'Downloads')  # Define o diretório de downloads
    arquivos = os.listdir(download_dir)  # Lista todos os arquivos no diretório de downloads

    # Filtra os arquivos para pegar apenas os arquivos CSV
    arquivos = [os.path.join(download_dir, f) for f in arquivos if f.endswith('.csv')]  
    if arquivos:  # Verifica se há arquivos CSV na pasta
        arquivo_recente = max(arquivos, key=os.path.getctime)  # Encontra o arquivo CSV mais recente

        # Define o novo nome e caminho para salvar o arquivo Excel
        data_atual = datetime.now().strftime("%d.%m")  # Obtém a data atual no formato 'dd.mm'
        novo_nome = f"Planejado - {data_atual}.xlsx"  # Formata o nome do novo arquivo
        destino = f"H:/Monitoramento_e_Avaliacao/Relatórios de Metas/Mensal/Planejado/{novo_nome}"  # Define o caminho de destino

        # Carrega o arquivo CSV em um DataFrame
        df = pd.read_csv(arquivo_recente, delimiter=";")  # Lê o arquivo CSV com ponto e vírgula como delimitador

        # Cria a coluna 'AREA2' concatenando 'AREA' e 'ID'
        if 'AREA' in df.columns and 'ID' in df.columns:
            df['AREA2'] = df['AREA'].astype(str) + '-' + df['ID'].astype(str)  # Cria a coluna 'AREA2' como string concatenada

        # Converte a coluna 'Data_Hora' para o formato datetime, se existir
        if 'Data_Hora' in df.columns:
            df['Data_Hora'] = pd.to_datetime(df['Data_Hora'], errors='coerce')  # Converte a coluna 'Data_Hora' para datetime

            # Cria as colunas 'Data' e 'Hora' separadas
            df['Data'] = df['Data_Hora'].dt.date  # Extrai a data
            df['Hora'] = df['Data_Hora'].dt.time  # Extrai a hora

            # Exclui a coluna original 'Data_Hora'
            df.drop(columns=['Data_Hora'], inplace=True)

            # Reposiciona as colunas 'Data' e 'Hora'
            colunas = list(df.columns)
            colunas.insert(6, colunas.pop(colunas.index('Data')))  # Move a coluna 'Data' para a posição 6
            colunas.insert(7, colunas.pop(colunas.index('Hora')))  # Move a coluna 'Hora' para a posição 7
            colunas.insert(2, colunas.pop(colunas.index('AREA')))  # Move a coluna 'AREA' para a posição 2
            df = df[colunas]  # Reorganiza as colunas conforme as modificações

        # Identifica o arquivo Excel mais recente na pasta de destino
        arquivos_excel = [os.path.join(os.path.dirname(destino), f) for f in os.listdir(os.path.dirname(destino)) if f.endswith('.xlsx')]
        if len(arquivos_excel) > 1:  # Verifica se existem mais de um arquivo Excel na pasta
            arquivos_excel.sort(key=os.path.getctime, reverse=True)  # Ordena os arquivos por data de criação
            mais_recente = arquivos_excel[0]  # Seleciona o arquivo mais recente
            df_antiga = pd.read_excel(mais_recente)  # Carrega o arquivo Excel mais recente

            # Realiza a operação de PROC V para preencher a coluna 'STATUS2'
            if 'AREA2' in df_antiga.columns and 'STATUS2' in df_antiga.columns:
                df_antiga['AREA2'] = df_antiga['AREA2'].astype(str)  # Garante que 'AREA2' esteja no formato correto
                df['STATUS2'] = df['AREA2'].map(df_antiga.set_index('AREA2')['STATUS2']).fillna('#N/D')  # Mapeia 'STATUS2' a partir do arquivo antigo
                df['STATUS2'] = df['STATUS2'].replace('0', '#N/D')  # Substitui valores '0' por '#N/D'

        # Salva o DataFrame como um arquivo Excel
        df.to_excel(destino, index=False)  # Salva o DataFrame como um arquivo Excel sem índice
        print(f"Arquivo '{novo_nome}' gerado com sucesso!")  # Exibe uma mensagem de sucesso
    else:
        print("Nenhum arquivo CSV encontrado na pasta de downloads.")  # Mensagem caso não haja arquivos CSV

# Função principal
def main():
    driver = configurar_driver()  # Configura o driver do navegador
    realizar_login(driver, config.USERNAME, config.PASSWORD)  # Realiza o login usando credenciais do arquivo de configurações
    navegar_painel(driver, config.USERNAME, config.PASSWORD)  # Navega até o painel após o login
    renomear_formatar_arquivo()  # Renomeia e formata o arquivo após a execução
    driver.quit()  # Encerra a sessão do navegador

if __name__ == "__main__":  # Verifica se o script está sendo executado diretamente
    main()  # Executa a função principal