import os
import time
from datetime import datetime
import pandas as pd
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
import config

# Função para configurar o driver
def configurar_driver():
    edge_options = Options()
    edge_service = EdgeService(executable_path=config.EDGE_DRIVER_PATH)
    driver = webdriver.Edge(service=edge_service, options=edge_options)
    driver.maximize_window()
    return driver

# Função para realizar o login
def realizar_login(driver, username, password):
    driver.get('https://suporte.santamarcelinacultura.org.br/planejamento/front/central.php') #Acessa o Site
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/form/div/div[1]/div[2]/input').send_keys(username) #Coloca o usuário do SPM
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/form/div/div[1]/div[3]/input').send_keys(password) #Coloca a senha do SPM
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/form/div/div[1]/div[5]/button').click() #Clica em entrar
    time.sleep(3) #Espera 3 segundos para a página carregar e para simular comportamento humano

# Função para navegar e realizar ações
def navegar_painel(driver, username, password):
    driver.find_element(By.XPATH, '/html/body/div[2]/header/div/div[2]/ul/li[1]/a/span').click() #Clica em ativos
    time.sleep(2) #Espera 2 segundos para simular comportamento humano
    driver.find_element(By.XPATH, '/html/body/div[2]/header/div/div[2]/ul/li[1]/div/div/div[2]/a[5]').click() #Clica em painel
    time.sleep(2) #Espera 2 segundos para simular comportamento humano
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/main/table/tbody/tr[2]/td/div[1]/a').click() #Clica em acessar painel
    time.sleep(3) #Espera 3 segundos para simular comportamento humano
    driver.find_element(By.XPATH, '/html/body/div/div/div[2]/form/div[1]/input').send_keys(username) #Coloca novamente o usuário do SPM
    driver.find_element(By.XPATH, '/html/body/div/div/div[2]/form/div[2]/input').send_keys(password) #Coloca novamente a senha do SPM 
    driver.find_element(By.XPATH, '/html/body/div/div/div[2]/form/button').click() #Clica em entrar
    time.sleep(3) #Espera 2 segundos para simular comportamento humano
    driver.find_element(By.XPATH, '/html/body/nav/button[5]').click() #Clica em "Download Realizado"

# Função para adicionar bordas em todas as células de uma planilha
def adicionar_bordas(sheet):
    thin_border = Border(  # Cria uma borda fina para as células
        left=Side(border_style="thin", color="000000"),  # Define a borda esquerda
        right=Side(border_style="thin", color="000000"),  # Define a borda direita
        top=Side(border_style="thin", color="000000"),  # Define a borda superior
        bottom=Side(border_style="thin", color="000000")  # Define a borda inferior
    )
    
    for row in sheet.iter_rows():  # Itera sobre todas as linhas da planilha
        for cell in row:  # Itera sobre todas as células da linha
            cell.border = thin_border  # Aplica a borda fina a cada célula

# Função para renomear, formatar e adicionar as colunas 'AREA2' e 'STATUS2' 
def renomear_formatar_arquivo():
    # Pasta de downloads (padrão) - o python acessa a pasta de downloads para localizar o arquivo realizado baixado
    download_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
    arquivos = os.listdir(download_dir)

    # Encontrar o arquivo CSV mais recente, que é justamente o realizado que acabamos de baixar
    arquivos_csv = [os.path.join(download_dir, f) for f in arquivos if f.endswith('.csv')]
    if arquivos_csv:
        arquivo_recente = max(arquivos_csv, key=os.path.getctime) 

        # Caminho para salvar o arquivo Excel
        data_atual = datetime.now().strftime("%d.%m") #Localiza a data atual
        novo_nome = f"Realizado - {data_atual}.xlsx" #Nomeia o arquivo com o padrão "Realizado - Data Atual"
        destino = f"H:/Monitoramento_e_Avaliacao/Relatórios de Metas/Mensal/Realizado/2024/{novo_nome}" #Salva o arquivo nomeado na pasta respectiva
        
        
##### O CÓDIGO ABAIXO TRATA DAS CUSTOMIZAÇÕES DA PLANILHA, ASSIM COMO DA REALIZAÇÃO DO PROCV #####       
        # Carregar o CSV em um DataFrame
        df = pd.read_csv(arquivo_recente, delimiter=";")  # O delimitador do .csv, neste caso, é o ponto e vírgula (';')

        # Criar a coluna 'AREA2'
        if 'AREA' in df.columns and 'ID' in df.columns:
            df['AREA2'] = df['AREA'].astype(str) + '-' + df['ID'].astype(str) #Aqui o Python cria a AREA2 a partir da concatenação de AREA e ID
        
        # Verificar e converter a coluna 'Data_Hora' para o tipo datetime
        if 'Data_Hora' in df.columns:
            df['Data_Hora'] = pd.to_datetime(df['Data_Hora'], errors='coerce')

            # Criar as colunas 'Data' e 'Hora' a partir da coluna 'Data_Hora'
            df['Data'] = df['Data_Hora'].dt.date  # Formatar no estilo DD/MM/YYYY
            df['Hora'] = df['Data_Hora'].dt.time

            # Excluir a coluna original 'Data_Hora', mantendo as colunas "Data" e "Hora" separadas
            df.drop(columns=['Data_Hora'], inplace=True)

            # Reposicionar as colunas 'Data' e 'Hora' para as posições de acordo com Metas_MAPA de Preenchimento
            colunas = list(df.columns)
            colunas.insert(6, colunas.pop(colunas.index('Data')))
            colunas.insert(7, colunas.pop(colunas.index('Hora')))
            colunas.insert(2, colunas.pop(colunas.index('AREA')))
            df = df[colunas]

        # Identificar o arquivo Excel mais recente na pasta de destino (M&A/RELATÓRIO DE METAS/MENSAL/REALIZADO)
        arquivos_excel = [os.path.join(os.path.dirname(destino), f) for f in os.listdir(os.path.dirname(destino)) if f.endswith('.xlsx')]
        if len(arquivos_excel) >= 2:
            arquivos_excel.sort(key=os.path.getctime, reverse=True)
            caminho_excel_antigo = arquivos_excel[0]
            print(f'Esse é o caminho do arquivo que está sendo analisado: {caminho_excel_antigo}')
            # Carregar o arquivo Excel para comparação
            df_antigo = pd.read_excel(caminho_excel_antigo)

            # Criar a coluna 'AREA2' no DataFrame antigo, se necessário
            if 'AREA' in df_antigo.columns and 'ID' in df_antigo.columns:
                df_antigo['AREA2'] = df_antigo['AREA'].astype(str) + '-' + df_antigo['ID'].astype(str)
                print('AREA2 CRIADA NA PLANILHA DE COMPARAÇÃO') #caso a planilha anterior já tenha a area2, e provavelmente terá, ele só vai sobrescrever a informação
            
            # Verificar e exibir os valores repetidos na coluna 'AREA2' do df_antigo - aqui somente verificamos se tem algum valor repetido em AREA2, o que não deve acontecer - se tiver, o terminal mostrará um erro
            valores_repetidos_df_antigo = df_antigo['AREA2'][df_antigo['AREA2'].duplicated(keep=False)]
            if not valores_repetidos_df_antigo.empty:
                print("Valores repetidos na coluna 'AREA2' do df_antigo:")
                print(valores_repetidos_df_antigo)

            # Aqui fazemos o procv da planilha antiga (o realizado mais recente sem ser o que acabou de ser baixado) e a nova (realizado que acabou de ser baixado)
            try:
                df['STATUS2'] = df['AREA2'].map(df_antigo.set_index('AREA2')['STATUS2']).fillna('#N/D') #se não houver correspondência, coloca #N/D. se houver, puxa o status que estava na planilha antiga
                print("Mapeamento realizado com sucesso.")
            except Exception as e:
                print(f"Erro durante o mapeamento: {e}") #Caso dê algum erro, o terminal mostrará essa mensagem
        else:
            print("Coluna 'STATUS2' não encontrada em df_antiga.") #Caso o usuário não deixe o realizado anterior como o arquivo mais recente na pasta de M&A, por exemplo, baixando os indicadores antes do realizado, o python não encontrará a coluna STATUS2 e retornará esse erro
            # Verificar colunas disponíveis em df_antiga
            print(f"Arquivo mais recente encontrado: {caminho_excel_antigo}") #Aqui ele mostra qual arquivo ele está considerando para fazer o procv (verifique se é o arquivo que faz sentido para o caso)
            print(f"Colunas disponíveis em df_antiga: {df_antigo.columns}")

        # Substituir todas as células vazias por 'N/A'
        df = df.fillna('N/A')

        # Salvar como Excel
        df.to_excel(destino, index=False)
        
        # Aplicar formatação usando openpyxl
        workbook = load_workbook(destino)
        sheet = workbook.active
        
        # Formatar cabeçalho
        font_bold = Font(bold=True)
        fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        for cell in sheet[1]:
            cell.font = font_bold
            cell.fill = fill_blue
        
        # Adicionar bordas
        adicionar_bordas(sheet) #aqui chamamos a função lá de cima para adicionar bordas
        
        # Definir altura das linhas e largura das colunas
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = 20

        # Ajustar largura de todas as colunas para 30
        for col in sheet.columns:
            column = col[0].column_letter  # Coluna atual
            sheet.column_dimensions[column].width = 30
        
        # Salvar o arquivo formatado
        workbook.save(destino)
        workbook.close()
        
        print(f"Arquivo renomeado, formatado e movido para: {destino}")
    else:
        print("Nenhum arquivo CSV encontrado na pasta de Downloads.")


# Função principal - aqui chamamos todos os códigos anterior em uma função geral, para rodar tudo
def tarefa():
    driver = configurar_driver()
    realizar_login(driver, config.USERNAME, config.PASSWORD)
    navegar_painel(driver, config.USERNAME, config.PASSWORD)
    time.sleep(5)
    driver.quit()
    renomear_formatar_arquivo()

if __name__ == "__main__":
    tarefa()
