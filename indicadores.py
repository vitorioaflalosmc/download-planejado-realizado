# Importação das bibliotecas necessárias
import os  # Para manipulação de arquivos e diretórios
import time  # Para manipulação de tempo e pausas no código
from datetime import datetime  # Para manipulação de datas e horários
import pandas as pd  # Para trabalhar com DataFrames e manipulação de dados em arquivos CSV
from selenium import webdriver  # Para automação do navegador web (Selenium)
from selenium.webdriver.edge.service import Service as EdgeService  # Para configurar o serviço do driver do Edge
from selenium.webdriver.edge.options import Options  # Para configurar opções do navegador Edge
from selenium.webdriver.common.by import By  # Para localizar elementos na página web
from openpyxl import load_workbook  # Para trabalhar com arquivos Excel (XLSX)
from openpyxl.styles import Font, PatternFill, Border, Side  # Para aplicar estilos no Excel (fontes, cores, bordas)
import config  # Arquivo de configuração para armazenar informações sensíveis, como o caminho do driver e credenciais

# Função para configurar o driver do navegador Edge
def configurar_driver():
    edge_options = Options()  # Cria uma instância das opções do navegador
    edge_service = EdgeService(executable_path=config.EDGE_DRIVER_PATH)  # Define o caminho do driver do Edge a partir do arquivo de configuração
    driver = webdriver.Edge(service=edge_service, options=edge_options)  # Inicializa o driver com as opções e o serviço definidos
    driver.maximize_window()  # Maximiza a janela do navegador
    return driver  # Retorna o driver configurado

# Função para realizar o login na página
def realizar_login(driver, username, password):
    driver.get('https://suporte.santamarcelinacultura.org.br/planejamento/front/central.php')  # Acessa a URL do SPM
    # Preenche o campo de nome de usuário com o valor recebido como parâmetro
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/form/div/div[1]/div[2]/input').send_keys(username)
    # Preenche o campo de senha com o valor recebido como parâmetro
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/form/div/div[1]/div[3]/input').send_keys(password)
    # Clica no botão de login para enviar as credenciais
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/form/div/div[1]/div[5]/button').click()
    time.sleep(3)  # Pausa de 3 segundos para garantir que a página carregue após o login

# Função para navegar pelo painel e realizar ações
def navegar_painel(driver, username, password):
    # Clica na opção de menu desejada para acessar a próxima tela
    driver.find_element(By.XPATH, '/html/body/div[2]/header/div/div[2]/ul/li[1]/a/span').click()
    time.sleep(2)  # Pausa de 2 segundos
    # Clica em outro item do menu para acessar a página específica
    driver.find_element(By.XPATH, '/html/body/div[2]/header/div/div[2]/ul/li[1]/div/div/div[2]/a[5]').click()
    time.sleep(2)  # Pausa de 2 segundos
    # Clica no primeiro item da lista de indicadores
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/main/table/tbody/tr[2]/td/div[1]/a').click()
    time.sleep(3)  # Pausa de 3 segundos
    # Preenche os campos de login na página de destino (repetição de login)
    driver.find_element(By.XPATH, '/html/body/div/div/div[2]/form/div[1]/input').send_keys(username)
    driver.find_element(By.XPATH, '/html/body/div/div/div[2]/form/div[2]/input').send_keys(password)
    # Clica no botão de confirmação para entrar
    driver.find_element(By.XPATH, '/html/body/div/div/div[2]/form/button').click()
    time.sleep(3)  # Pausa de 3 segundos para garantir que a página carregue

# Função para tentar converter valores para float
def tentar_converter_para_float(valor):
    try:
        return float(valor)  # Tenta converter o valor para float
    except ValueError:
        return valor  # Se não for possível, retorna o valor original

# Função para baixar e formatar o arquivo de indicadores
def baixar_formatar_arquivo_indicadores(username, password):
    driver = configurar_driver()  # Configura o driver do navegador
    realizar_login(driver, username, password)  # Realiza o login no site
    navegar_painel(driver, username, password)  # Navega pelo painel após o login

    # Clica no botão para iniciar o download do arquivo
    driver.find_element(By.XPATH, '/html/body/nav/button[3]').click()
    time.sleep(5)  # Pausa de 5 segundos para garantir que o clique tenha efeito
    # Clica no botão para confirmar o download do arquivo
    driver.find_element(By.XPATH, '/html/body/div[3]/button[4]').click()
    
    # Espera o download ser completado
    time.sleep(10)  # Pausa de 10 segundos para garantir que o download seja concluído
    driver.quit()  # Fecha o navegador após a conclusão da tarefa

    # Obtém o diretório padrão de downloads do sistema
    download_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
    arquivos = os.listdir(download_dir)  # Lista os arquivos presentes na pasta de downloads

    # Filtra para encontrar arquivos CSV
    arquivos = [os.path.join(download_dir, f) for f in arquivos if f.endswith('.csv')]
    if arquivos:  # Se existir algum arquivo CSV
        # Encontra o arquivo CSV mais recente
        arquivo_recente = max(arquivos, key=os.path.getctime)

        # Define o caminho onde o arquivo Excel será salvo
        data_atual = datetime.now().strftime("%d.%m")  # Formata a data atual
        destino = f"H:/Monitoramento_e_Avaliacao/Relatórios de Metas/Mensal/Realizado/2024/Indicadores - {data_atual}.xlsx"
        
        # Carrega o arquivo CSV em um DataFrame do pandas
        df = pd.read_csv(arquivo_recente, delimiter=";")

        # Remove a coluna 'planejado' se ela existir no DataFrame
        if 'planejado' in df.columns:
            df = df.drop(columns=['planejado'])

        # Tenta converter a coluna 'realizado' para valores float
        df['realizado'] = df['realizado'].apply(tentar_converter_para_float)

        # Remove linhas onde a coluna 'realizado' tem valor 0
        df = df[~((df['realizado'] == 0) & (df['realizado'].apply(lambda x: isinstance(x, float))))]

        # Salva o DataFrame como um arquivo Excel
        df.to_excel(destino, index=False)
        
        # Carrega o arquivo Excel para formatar
        workbook = load_workbook(destino)
        sheet = workbook.active
        
        # Aplica formatação ao cabeçalho (negrito e cor de fundo)
        font_bold = Font(bold=True)
        fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for cell in sheet[1]:
            cell.font = font_bold
            cell.fill = fill_blue
        
        # Adiciona bordas finas nas células
        thin_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Define a altura das linhas
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = 20

        # Ajusta a largura das colunas para 30
        for col in sheet.columns:
            column = col[0].column_letter  # Obtém a letra da coluna
            sheet.column_dimensions[column].width = 30
        
        # Salva o arquivo Excel com as formatações aplicadas
        workbook.save(destino)
        workbook.close()  # Fecha o arquivo Excel
        
        print(f"Arquivo renomeado, formatado e movido para: {destino}")  # Exibe uma mensagem de sucesso
    else:
        print("Nenhum arquivo CSV encontrado na pasta de Downloads.")  # Exibe uma mensagem se nenhum arquivo for encontrado

# Bloco principal, executa a função de baixar e formatar o arquivo de indicadores
if __name__ == "__main__":
    username = config.USERNAME  # Carrega o nome de usuário da configuração
    password = config.PASSWORD  # Carrega a senha da configuração
    baixar_formatar_arquivo_indicadores(username, password)  # Chama a função para realizar as ações
